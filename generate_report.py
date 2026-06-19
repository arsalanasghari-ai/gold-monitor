"""
generate_report.py
ساخت گزارش اکسل از فایل price_archive.csv
نوع گزارش: weekly / monthly / yearly
"""

import sys
import csv
import os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

ARCHIVE_FILE = "price_archive.csv"
REPORT_TYPE = os.environ.get("REPORT_TYPE", "weekly")  # weekly / monthly / yearly


# ===================== خواندن و فیلتر داده =====================

def read_archive():
    if not os.path.exists(ARCHIVE_FILE):
        print(f"❌ فایل {ARCHIVE_FILE} پیدا نشد")
        return []
    rows = []
    with open(ARCHIVE_FILE, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                row['datetime'] = datetime.strptime(
                    f"{row['date']} {row['time']}", "%Y-%m-%d %H:%M"
                )
                rows.append(row)
            except Exception:
                continue
    return rows

def filter_rows(rows, report_type):
    now = datetime.now()
    if report_type == "weekly":
        since = now - timedelta(days=7)
    elif report_type == "monthly":
        since = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif report_type == "yearly":
        since = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        since = now - timedelta(days=7)
    return [r for r in rows if r['datetime'] >= since]

def daily_summary(rows):
    """خلاصه روزانه: اول، آخر، بیشترین، کمترین"""
    by_day = {}
    for r in rows:
        d = r['date']
        if d not in by_day:
            by_day[d] = []
        by_day[d].append(r)

    summary = []
    for d in sorted(by_day.keys()):
        day_rows = by_day[d]
        def safe_float(row, key):
            try:
                return float(row.get(key, 0) or 0)
            except Exception:
                return 0.0

        gold_vals = [safe_float(r, 'gold_18k_rial') for r in day_rows if safe_float(r, 'gold_18k_rial') > 0]
        btc_vals  = [safe_float(r, 'bitcoin_usd')   for r in day_rows if safe_float(r, 'bitcoin_usd') > 0]
        tether_vals = [safe_float(r, 'tether_rial') for r in day_rows if safe_float(r, 'tether_rial') > 0]
        ounce_vals  = [safe_float(r, 'gold_ounce_usd') for r in day_rows if safe_float(r, 'gold_ounce_usd') > 0]

        summary.append({
            'date': d,
            'gold_open':   gold_vals[0]  if gold_vals   else 0,
            'gold_close':  gold_vals[-1] if gold_vals   else 0,
            'gold_high':   max(gold_vals) if gold_vals  else 0,
            'gold_low':    min(gold_vals) if gold_vals  else 0,
            'btc_open':    btc_vals[0]   if btc_vals    else 0,
            'btc_close':   btc_vals[-1]  if btc_vals    else 0,
            'btc_high':    max(btc_vals) if btc_vals    else 0,
            'btc_low':     min(btc_vals) if btc_vals    else 0,
            'tether_close': tether_vals[-1] if tether_vals else 0,
            'ounce_close':  ounce_vals[-1]  if ounce_vals  else 0,
        })
    return summary


# ===================== استایل‌ها =====================

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
SUBHDR_FILL   = PatternFill("solid", start_color="2E75B6")
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL    = PatternFill("solid", start_color="FFFFFF")
GREEN_FILL    = PatternFill("solid", start_color="C6EFCE")
RED_FILL      = PatternFill("solid", start_color="FFC7CE")

HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
NORMAL_FONT   = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", bold=True, size=10)
TITLE_FONT    = Font(name="Arial", bold=True, size=14, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, sub=False):
    cell.fill   = SUBHDR_FILL if sub else HEADER_FILL
    cell.font   = SUBHDR_FONT if sub else HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_cell(cell, alt=False, bold=False, align=RIGHT):
    cell.fill      = ALT_FILL if alt else WHITE_FILL
    cell.font      = BOLD_FONT if bold else NORMAL_FONT
    cell.alignment = align
    cell.border    = BORDER

def pct_color(ws, cell_ref, base_val, cur_val):
    cell = ws[cell_ref]
    if base_val and base_val > 0:
        pct = (cur_val - base_val) / base_val
        if pct > 0.005:
            cell.fill = GREEN_FILL
        elif pct < -0.005:
            cell.fill = RED_FILL


# ===================== شیت ۱: خلاصه روزانه =====================

def write_daily_sheet(wb, summary, report_type):
    ws = wb.active
    ws.title = "خلاصه روزانه"
    ws.sheet_view.rightToLeft = True

    period_fa = {"weekly": "هفتگی", "monthly": "ماهانه", "yearly": "سالانه"}.get(report_type, "")
    ws.merge_cells("A1:K1")
    ws["A1"] = f"📊 گزارش {period_fa} بازار مالی — {datetime.now().strftime('%Y-%m-%d')}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].fill = PatternFill("solid", start_color="EBF3FB")
    ws.row_dimensions[1].height = 28

    headers = [
        "تاریخ",
        "طلا ۱۸ع — باز (ریال)", "طلا ۱۸ع — بسته (ریال)",
        "طلا ۱۸ع — سقف",        "طلا ۱۸ع — کف",
        "تغییر طلا %",
        "بیت‌کوین باز ($)", "بیت‌کوین بسته ($)",
        "تغییر BTC %",
        "تتر (ریال)",
        "اونس جهانی ($)"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c)
    ws.row_dimensions[2].height = 22

    for i, row in enumerate(summary):
        r = i + 3
        alt = (i % 2 == 1)
        data = [
            row['date'],
            row['gold_open']  or None,
            row['gold_close'] or None,
            row['gold_high']  or None,
            row['gold_low']   or None,
            f"={(get_column_letter(3)}{r}-{get_column_letter(2)}{r})/{get_column_letter(2)}{r}" if row['gold_open'] else None,
            row['btc_open']   or None,
            row['btc_close']  or None,
            f"={(get_column_letter(8)}{r}-{get_column_letter(7)}{r})/{get_column_letter(7)}{r}" if row['btc_open'] else None,
            row['tether_close'] or None,
            row['ounce_close']  or None,
        ]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=val)
            style_cell(c, alt=alt, align=CENTER if col == 1 else RIGHT)
            if col in (6, 9) and val:
                c.number_format = "0.00%"
            elif col in (2, 3, 4, 5, 10):
                c.number_format = "#,##0"
            elif col in (7, 8, 11):
                c.number_format = "#,##0.00"

        if row['gold_open']:
            pct_color(ws, f"F{r}", row['gold_open'], row['gold_close'])
        if row['btc_open']:
            pct_color(ws, f"I{r}", row['btc_open'], row['btc_close'])

    # سطر جمع / میانگین
    last_data = 2 + len(summary)
    sum_row = last_data + 1
    ws.cell(row=sum_row, column=1, value="میانگین دوره").font = BOLD_FONT
    ws.cell(row=sum_row, column=1).alignment = CENTER
    for col_idx, col_letter in [(3, "C"), (8, "H"), (10, "J"), (11, "K")]:
        formula = f"=IFERROR(AVERAGE({col_letter}3:{col_letter}{last_data}),\"\")"
        c = ws.cell(row=sum_row, column=col_idx, value=formula)
        c.font = BOLD_FONT
        c.fill = PatternFill("solid", start_color="D9E1F2")
        c.alignment = RIGHT
        c.border = BORDER
        if col_idx in (3, 10):
            c.number_format = "#,##0"
        else:
            c.number_format = "#,##0.00"

    col_widths = [12, 20, 20, 18, 18, 12, 18, 18, 12, 18, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws, last_data


# ===================== شیت ۲: میانگین متحرک =====================

def moving_average(values, window):
    result = []
    for i in range(len(values)):
        if i < window - 1:
            result.append(None)
        else:
            result.append(sum(values[i-window+1:i+1]) / window)
    return result

def write_ma_sheet(wb, summary):
    ws = wb.create_sheet("میانگین متحرک")
    ws.sheet_view.rightToLeft = True

    ws.merge_cells("A1:H1")
    ws["A1"] = "📈 میانگین متحرک (MA7 و MA14)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws["A1"].fill = PatternFill("solid", start_color="EBF3FB")
    ws.row_dimensions[1].height = 26

    headers = ["تاریخ", "طلا بسته", "MA7 طلا", "MA14 طلا",
               "BTC بسته", "MA7 BTC", "MA14 BTC", "اونس بسته"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        style_header(c, sub=True)

    gold_close = [r['gold_close'] for r in summary]
    btc_close  = [r['btc_close']  for r in summary]
    ma7_gold   = moving_average(gold_close, 7)
    ma14_gold  = moving_average(gold_close, 14)
    ma7_btc    = moving_average(btc_close, 7)
    ma14_btc   = moving_average(btc_close, 14)

    for i, row in enumerate(summary):
        r = i + 3
        alt = (i % 2 == 1)
        data = [
            row['date'],
            gold_close[i]  or None,
            ma7_gold[i],
            ma14_gold[i],
            btc_close[i]   or None,
            ma7_btc[i],
            ma14_btc[i],
            row['ounce_close'] or None,
        ]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=r, column=col, value=round(val, 2) if val else None)
            style_cell(c, alt=alt, align=CENTER if col == 1 else RIGHT)
            if col in (2, 3, 4):
                c.number_format = "#,##0"
            elif col in (5, 6, 7, 8):
                c.number_format = "#,##0.00"

    col_widths = [12, 18, 14, 14, 16, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # نمودار میانگین طلا
    if len(summary) >= 2:
        chart = LineChart()
        chart.title = "طلای ۱۸ عیار + میانگین متحرک"
        chart.style = 10
        chart.y_axis.title = "قیمت (ریال)"
        chart.x_axis.title = "تاریخ"
        chart.width = 22
        chart.height = 12

        last_r = 2 + len(summary)
        cats = Reference(ws, min_col=1, min_row=3, max_row=last_r)
        for col_idx, title in [(2, "طلا بسته"), (3, "MA7"), (4, "MA14")]:
            data_ref = Reference(ws, min_col=col_idx, min_row=2, max_row=last_r)
            chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "A" + str(last_r + 3))

    return ws


# ===================== شیت ۳: داده کامل ساعتی =====================

def write_raw_sheet(wb, rows):
    ws = wb.create_sheet("داده ساعتی خام")
    ws.sheet_view.rightToLeft = True

    headers = ["تاریخ", "ساعت", "طلا ۱۸ع (ریال)", "اونس ($)", "BTC ($)", "تتر (ریال)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        style_header(c, sub=True)

    for i, row in enumerate(rows, 2):
        alt = (i % 2 == 0)
        data = [
            row.get('date', ''),
            row.get('time', ''),
            float(row.get('gold_18k_rial', 0) or 0) or None,
            float(row.get('gold_ounce_usd', 0) or 0) or None,
            float(row.get('bitcoin_usd', 0) or 0) or None,
            float(row.get('tether_rial', 0) or 0) or None,
        ]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=i, column=col, value=val)
            style_cell(c, alt=alt, align=CENTER if col <= 2 else RIGHT)
            if col == 3:
                c.number_format = "#,##0"
            elif col in (4, 5, 6):
                c.number_format = "#,##0.00"

    col_widths = [12, 8, 20, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return ws


# ===================== اجرای اصلی =====================

def main():
    report_type = REPORT_TYPE
    print(f"⏰ ساخت گزارش {report_type} — {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    all_rows = read_archive()
    if not all_rows:
        print("❌ داده‌ای در آرشیو پیدا نشد")
        sys.exit(1)

    filtered = filter_rows(all_rows, report_type)
    if len(filtered) < 2:
        print(f"⚠️ داده کافی برای گزارش {report_type} وجود ندارد ({len(filtered)} رکورد)")
        sys.exit(0)

    summary = daily_summary(filtered)
    print(f"✅ {len(filtered)} رکورد ساعتی، {len(summary)} روز")

    wb = Workbook()
    write_daily_sheet(wb, summary, report_type)
    write_ma_sheet(wb, summary)
    write_raw_sheet(wb, filtered)

    now = datetime.now()
    filename = f"market_report_{report_type}_{now.strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    print(f"✅ فایل ذخیره شد: {filename}")

if __name__ == "__main__":
    main()
